#!/usr/bin/env python
"""Export all Football Fam players to an Excel validation workbook.

Pulls every player from the live API (page by page), groups them by
pyramid step (1-6), and writes one worksheet per step with formatting
suitable for manual data-quality review.

Features:
  - 30 columns covering profile, stats, and metadata
  - Auto-filters on every column
  - Frozen header row
  - Alternate-row shading (light blue)
  - Red highlight on rows where the full name looks like a number
  - Summary sheet with field-completeness percentages per step
  - VALIDATION column for the reviewer to fill in

Usage::

    python scripts/export_validation.py
    python scripts/export_validation.py --full          # fetch per-player detail (slow)
    python scripts/export_validation.py --api http://localhost:8000/api/v1
    python scripts/export_validation.py --per-page 50
"""

import argparse
import logging
import re
import sys
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
)
log = logging.getLogger(__name__)

DEFAULT_API_BASE = (
    "https://football-fam-data-production.up.railway.app/api/v1"
)
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "data"
OUTPUT_FILE = OUTPUT_DIR / "Football_Fam_Validation_Export.xlsx"

COLUMNS = [
    "Player ID",
    "Full Name",
    "Position",
    "Position Detail",
    "Date of Birth",
    "Age",
    "Nationality",
    "Height (cm)",
    "Weight (kg)",
    "Preferred Foot",
    "Club",
    "League",
    "Step",
    "Contract Status",
    "Availability",
    "Appearances",
    "Starts",
    "Sub Appearances",
    "Goals",
    "Assists",
    "Yellow Cards",
    "Red Cards",
    "Clean Sheets",
    "Minutes Played",
    "Data Source",
    "Confidence",
    "Has Photo",
    "Photo URL",
    "Is Verified",
    "VALIDATION",
]

COLUMN_WIDTHS: dict[str, int] = {
    "Player ID": 10,
    "Full Name": 30,
    "Position": 10,
    "Position Detail": 16,
    "Date of Birth": 14,
    "Age": 6,
    "Nationality": 16,
    "Height (cm)": 11,
    "Weight (kg)": 11,
    "Preferred Foot": 13,
    "Club": 28,
    "League": 32,
    "Step": 6,
    "Contract Status": 16,
    "Availability": 15,
    "Appearances": 12,
    "Starts": 8,
    "Sub Appearances": 15,
    "Goals": 7,
    "Assists": 8,
    "Yellow Cards": 12,
    "Red Cards": 10,
    "Clean Sheets": 12,
    "Minutes Played": 14,
    "Data Source": 14,
    "Confidence": 11,
    "Has Photo": 10,
    "Photo URL": 30,
    "Is Verified": 11,
    "VALIDATION": 20,
}

# Fields checked for completeness in the summary sheet
COMPLETENESS_FIELDS: list[tuple[str, str]] = [
    ("DOB", "date_of_birth"),
    ("Nationality", "nationality"),
    ("Position", "position_primary"),
    ("Height", "height_cm"),
    ("Weight", "weight_kg"),
    ("Foot", "preferred_foot"),
    ("Club", "_club_name"),
    ("Photo", "profile_photo_url"),
    ("Stats", "_has_stats"),
    ("Verified", "is_verified"),
]


# ══════════════════════════════════════════════════════════════════════════
# API fetching
# ══════════════════════════════════════════════════════════════════════════

def fetch_all_players(
    api_base: str,
    per_page: int = 100,
) -> list[dict[str, Any]]:
    """Pull every player from the search endpoint, page by page."""
    all_players: list[dict[str, Any]] = []
    page = 1
    total_pages = None

    session = requests.Session()

    while True:
        url = f"{api_base}/players/search"
        params = {
            "sort_by": "name",
            "page": page,
            "per_page": per_page,
        }

        log.info(
            "Fetching page %s%s …",
            page,
            f"/{total_pages}" if total_pages else "",
        )

        try:
            resp = session.get(url, params=params, timeout=30)
            resp.raise_for_status()
        except requests.RequestException as exc:
            log.error("API request failed on page %d: %s", page, exc)
            break

        data = resp.json()
        results = data.get("results", [])
        total = data.get("total", 0)
        total_pages = data.get("pages", 1)

        if not results:
            break

        all_players.extend(results)
        log.info(
            "  Got %d players (total fetched: %d / %d)",
            len(results), len(all_players), total,
        )

        if page >= total_pages:
            break

        page += 1
        time.sleep(0.3)

    log.info("Fetched %d players in total", len(all_players))
    return all_players


def enrich_with_detail(
    players: list[dict[str, Any]],
    api_base: str,
) -> None:
    """Call /players/{id}/stats for each player to get data_source and sub_appearances.

    Mutates the player dicts in place, adding ``_detail_stats`` with
    the richer stat data from the individual endpoint.
    """
    session = requests.Session()
    total = len(players)

    for i, player in enumerate(players, start=1):
        pid = player.get("id")
        if not pid:
            continue

        if i % 100 == 0 or i == 1:
            log.info("  Enriching player %d / %d …", i, total)

        try:
            resp = session.get(
                f"{api_base}/players/{pid}/stats", timeout=15,
            )
            if resp.status_code == 200:
                stats_list = resp.json()
                if isinstance(stats_list, list) and stats_list:
                    best = stats_list[0]
                    player["_detail_stats"] = best
        except requests.RequestException:
            pass

        time.sleep(0.15)

    log.info("  Enrichment complete for %d players", total)


# ══════════════════════════════════════════════════════════════════════════
# Data helpers
# ══════════════════════════════════════════════════════════════════════════

def _get_step(player: dict[str, Any]) -> int:
    league = player.get("league")
    if league and league.get("step"):
        return int(league["step"])
    return 0


def _looks_like_number(name: str) -> bool:
    if not name:
        return True
    stripped = re.sub(r"[\s\-'.]", "", name)
    if not stripped:
        return True
    digit_ratio = sum(c.isdigit() for c in stripped) / len(stripped)
    return digit_ratio > 0.5


def _player_row(player: dict[str, Any]) -> list[Any]:
    """Convert a player API dict into a flat row matching COLUMNS."""
    club = player.get("club") or {}
    league = player.get("league") or {}
    stats = player.get("season_stats") or {}
    detail = player.get("_detail_stats") or {}

    has_photo = "Yes" if player.get("profile_photo_url") else "No"
    is_verified = "Yes" if player.get("is_verified") else "No"

    appearances = stats.get("appearances")
    starts = stats.get("starts")
    sub_apps = detail.get("sub_appearances")
    if sub_apps is None and appearances is not None and starts is not None:
        sub_apps = max(0, appearances - starts)

    data_source = detail.get("data_source") or stats.get("data_source", "")

    return [
        player.get("id"),
        player.get("full_name", ""),
        player.get("position_primary", ""),
        player.get("position_detail", ""),
        player.get("date_of_birth", ""),
        player.get("age"),
        player.get("nationality", ""),
        player.get("height_cm"),
        player.get("weight_kg"),
        player.get("preferred_foot", ""),
        club.get("name", ""),
        league.get("name", ""),
        league.get("step"),
        player.get("contract_status", ""),
        player.get("availability", ""),
        appearances,
        starts,
        sub_apps,
        stats.get("goals"),
        stats.get("assists"),
        stats.get("yellow_cards"),
        stats.get("red_cards"),
        stats.get("clean_sheets"),
        stats.get("minutes_played"),
        data_source,
        stats.get("confidence_score"),
        has_photo,
        player.get("profile_photo_url", ""),
        is_verified,
        "",
    ]


def _field_present(player: dict[str, Any], key: str) -> bool:
    """Check if a field has a meaningful value for completeness stats."""
    if key == "_club_name":
        club = player.get("club") or {}
        return bool(club.get("name"))
    if key == "_has_stats":
        return player.get("season_stats") is not None
    if key == "is_verified":
        return bool(player.get("is_verified"))
    val = player.get(key)
    if val is None or val == "" or val == "unknown":
        return False
    return True


# ══════════════════════════════════════════════════════════════════════════
# Excel styles
# ══════════════════════════════════════════════════════════════════════════

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
BAD_NAME_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BAD_NAME_FONT = Font(name="Calibri", color="9C0006", size=11)
NORMAL_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14)
PCT_GREEN = Font(name="Calibri", color="006100", size=11)
PCT_AMBER = Font(name="Calibri", color="9C6500", size=11)
PCT_RED = Font(name="Calibri", color="9C0006", size=11)
THIN_BORDER = Border(bottom=Side(style="thin", color="B0B0B0"))
HEADER_BORDER = Border(bottom=Side(style="medium", color="1F4E79"))

TAB_COLOURS: dict[int, str] = {
    1: "1F77B4", 2: "FF7F0E", 3: "2CA02C",
    4: "D62728", 5: "9467BD", 6: "8C564B",
    0: "999999",
}


# ══════════════════════════════════════════════════════════════════════════
# Workbook building
# ══════════════════════════════════════════════════════════════════════════

def build_step_sheets(
    wb: Workbook,
    players_by_step: dict[int, list[dict[str, Any]]],
) -> None:
    """Create one formatted worksheet per step with all player data."""
    for step in sorted(players_by_step.keys()):
        players = players_by_step[step]
        sheet_name = f"Step {step}" if step > 0 else "Unknown Step"
        ws = wb.create_sheet(title=sheet_name)

        for col_idx, header in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = HEADER_BORDER

        for row_idx, player in enumerate(players, start=2):
            values = _player_row(player)
            is_alt = (row_idx % 2) == 0
            name_is_bad = _looks_like_number(player.get("full_name", ""))

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")

                if name_is_bad:
                    cell.fill = BAD_NAME_FILL
                    cell.font = BAD_NAME_FONT
                elif is_alt:
                    cell.fill = ALT_ROW_FILL

        for col_idx, header in enumerate(COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = COLUMN_WIDTHS.get(header, 14)

        last_col = get_column_letter(len(COLUMNS))
        last_row = len(players) + 1
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"
        ws.freeze_panes = "A2"
        ws.sheet_properties.tabColor = TAB_COLOURS.get(step, "999999")

        log.info("  Sheet '%s': %d players", sheet_name, len(players))


def build_summary_sheet(
    wb: Workbook,
    players_by_step: dict[int, list[dict[str, Any]]],
    all_players: list[dict[str, Any]],
    api_base: str,
    full_mode: bool,
) -> None:
    """Create the Summary sheet at index 0 with counts and completeness."""
    ws = wb.create_sheet(title="Summary", index=0)
    ws.sheet_properties.tabColor = "333333"

    row = 1

    # ── Title block ───────────────────────────────────────────────
    ws.cell(row=row, column=1, value="Football Fam — Validation Export").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = NORMAL_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"API: {api_base}").font = NORMAL_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"Mode: {'full (per-player detail)' if full_mode else 'standard (search only)'}").font = NORMAL_FONT
    row += 2

    # ── Counts table ──────────────────────────────────────────────
    count_headers = ["Step", "Players", "With Photo", "With Stats", "Bad Names"]
    for col_idx, h in enumerate(count_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    row += 1

    for step in sorted(players_by_step.keys()):
        sp = players_by_step[step]
        label = f"Step {step}" if step > 0 else "Unknown"
        values = [
            label,
            len(sp),
            sum(1 for p in sp if p.get("profile_photo_url")),
            sum(1 for p in sp if p.get("season_stats")),
            sum(1 for p in sp if _looks_like_number(p.get("full_name", ""))),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = NORMAL_FONT
            if row % 2 == 0:
                cell.fill = ALT_ROW_FILL
        row += 1

    ws.cell(row=row, column=1, value="Total").font = BOLD_FONT
    ws.cell(row=row, column=2, value=len(all_players)).font = BOLD_FONT
    row += 2

    # ── Field completeness table ──────────────────────────────────
    ws.cell(row=row, column=1, value="Field Completeness (%)").font = TITLE_FONT
    row += 1

    comp_headers = ["Step"] + [label for label, _ in COMPLETENESS_FIELDS]
    for col_idx, h in enumerate(comp_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    row += 1

    def _pct_font(pct: float) -> Font:
        if pct >= 80:
            return PCT_GREEN
        if pct >= 50:
            return PCT_AMBER
        return PCT_RED

    for step in sorted(players_by_step.keys()):
        sp = players_by_step[step]
        label = f"Step {step}" if step > 0 else "Unknown"
        total = len(sp) or 1
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT

        for col_idx, (_, field_key) in enumerate(COMPLETENESS_FIELDS, start=2):
            present = sum(1 for p in sp if _field_present(p, field_key))
            pct = round(100 * present / total, 1)
            cell = ws.cell(row=row, column=col_idx, value=f"{pct}%")
            cell.font = _pct_font(pct)
            cell.alignment = Alignment(horizontal="center")

        if row % 2 == 0:
            for col_idx in range(1, len(comp_headers) + 1):
                existing_fill = ws.cell(row=row, column=col_idx).fill
                if existing_fill == PatternFill():
                    ws.cell(row=row, column=col_idx).fill = ALT_ROW_FILL
        row += 1

    # All players row
    ws.cell(row=row, column=1, value="All").font = BOLD_FONT
    total = len(all_players) or 1
    for col_idx, (_, field_key) in enumerate(COMPLETENESS_FIELDS, start=2):
        present = sum(1 for p in all_players if _field_present(p, field_key))
        pct = round(100 * present / total, 1)
        cell = ws.cell(row=row, column=col_idx, value=f"{pct}%")
        cell.font = _pct_font(pct)
        cell.alignment = Alignment(horizontal="center")
    row += 2

    # ── Data source breakdown ─────────────────────────────────────
    ws.cell(row=row, column=1, value="Data Sources").font = TITLE_FONT
    row += 1

    source_counts: dict[str, int] = defaultdict(int)
    for p in all_players:
        stats = p.get("season_stats") or {}
        detail = p.get("_detail_stats") or {}
        src = detail.get("data_source") or stats.get("data_source", "")
        if src:
            source_counts[src] += 1
        else:
            source_counts["(none)"] += 1

    for col_idx, h in enumerate(["Source", "Players"], start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    row += 1

    for src, count in sorted(source_counts.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=src).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=count).font = NORMAL_FONT
        row += 1

    # ── Column widths ─────────────────────────────────────────────
    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, len(comp_headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13
    ws.freeze_panes = "A1"


# ══════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export Football Fam players to Excel validation workbook",
    )
    parser.add_argument(
        "--api", type=str, default=DEFAULT_API_BASE,
        help=f"API base URL (default: {DEFAULT_API_BASE})",
    )
    parser.add_argument(
        "--per-page", type=int, default=100,
        help="Players per API page (default: 100)",
    )
    parser.add_argument(
        "--full", action="store_true",
        help="Fetch individual player stats for data_source and sub_appearances (slow)",
    )
    parser.add_argument(
        "--output", type=str, default=None,
        help="Output file path (default: data/Football_Fam_Validation_Export.xlsx)",
    )
    args = parser.parse_args()

    api_base = args.api.rstrip("/")
    output_path = Path(args.output) if args.output else OUTPUT_FILE

    log.info("=" * 60)
    log.info("Football Fam — Validation Export")
    log.info("  API:    %s", api_base)
    log.info("  Mode:   %s", "full (per-player detail)" if args.full else "standard")
    log.info("  Output: %s", output_path)
    log.info("=" * 60)

    # ── 1. Fetch all players ──────────────────────────────────────
    players = fetch_all_players(api_base, per_page=args.per_page)
    if not players:
        log.error("No players returned from API — nothing to export")
        sys.exit(1)

    # ── 2. Optional enrichment ────────────────────────────────────
    if args.full:
        log.info("Enriching %d players with per-player detail …", len(players))
        enrich_with_detail(players, api_base)

    # ── 3. Group by step ──────────────────────────────────────────
    by_step: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for p in players:
        by_step[_get_step(p)].append(p)

    log.info("Players grouped by step:")
    for step in sorted(by_step.keys()):
        label = f"Step {step}" if step > 0 else "Unknown"
        log.info("  %s: %d players", label, len(by_step[step]))

    # ── 4. Build workbook ─────────────────────────────────────────
    log.info("Building Excel workbook …")
    wb = Workbook()
    wb.remove(wb.active)

    build_step_sheets(wb, by_step)
    build_summary_sheet(wb, by_step, players, api_base, args.full)

    # ── 5. Save ───────────────────────────────────────────────────
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    bad_total = sum(
        1 for p in players if _looks_like_number(p.get("full_name", ""))
    )

    log.info("=" * 60)
    log.info("  Export complete!")
    log.info("  File:          %s", output_path)
    log.info("  Total players: %d", len(players))
    log.info("  Sheets:        %d step sheets + 1 summary", len(by_step))
    if bad_total:
        log.warning("  Flagged names: %d (highlighted in red)", bad_total)
    log.info("=" * 60)


if __name__ == "__main__":
    main()
