#!/usr/bin/env python3
"""Export pending merge candidates to an Excel workbook for agent review.

For each candidate pair the sheet shows Player A and Player B details
side-by-side, the match score, a breakdown of match reasons, and a
yellow DECISION column for the reviewer to fill in (MERGE or REJECT).

Sorted by score descending so the highest-confidence duplicates come first.

Usage::

    python scripts/export_merge_candidates.py
"""

import logging
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.orm import aliased

from src.db.models import (
    Club,
    League,
    MergeCandidate,
    Player,
    PlayerSeason,
)
from src.db.session import get_session

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
)
log = logging.getLogger(__name__)

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "data"
OUTPUT_FILE = OUTPUT_DIR / "Merge_Candidates_Review.xlsx"

# ── Styles (consistent with export_validation.py) ────────────────────────

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
NORMAL_FONT = Font(name="Calibri", size=11)
ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DECISION_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
DECISION_FONT = Font(name="Calibri", bold=True, size=11)
THIN_BORDER = Border(bottom=Side(style="thin", color="B0B0B0"))
HEADER_BORDER = Border(bottom=Side(style="medium", color="1F4E79"))

COLUMNS = [
    "Candidate ID",
    "Score",
    # Player A
    "A: Player ID",
    "A: Name",
    "A: DOB",
    "A: Position",
    "A: Club",
    "A: League",
    "A: Step",
    "A: Appearances",
    "A: Goals",
    # Player B
    "B: Player ID",
    "B: Name",
    "B: DOB",
    "B: Position",
    "B: Club",
    "B: League",
    "B: Step",
    "B: Appearances",
    "B: Goals",
    # Reasons
    "Name Similarity",
    "Same Club",
    "Same Position",
    "DOB Match",
    "Same Nationality",
    "Other Reasons",
    # Decision
    "DECISION",
]


def _age(dob: date | None) -> int | None:
    if not dob:
        return None
    today = date.today()
    return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))


def _get_total_stats(session, player_id: int) -> tuple[int, int]:
    """Sum appearances and goals across all seasons for a player."""
    row = session.execute(
        select(
            func.coalesce(func.sum(PlayerSeason.appearances), 0),
            func.coalesce(func.sum(PlayerSeason.goals), 0),
        ).where(PlayerSeason.player_id == player_id)
    ).one()
    return int(row[0]), int(row[1])


def _player_info(session, player: Player) -> dict:
    """Build a flat dict of the fields we need for a player."""
    club_name = ""
    league_name = ""
    step = None

    if player.current_club_id:
        club = session.get(Club, player.current_club_id)
        if club:
            club_name = club.name
            if club.league_id:
                league = session.get(League, club.league_id)
                if league:
                    league_name = league.name
                    step = league.step

    apps, goals = _get_total_stats(session, player.id)

    return {
        "id": player.id,
        "name": player.full_name,
        "dob": player.date_of_birth.isoformat() if player.date_of_birth else "",
        "position": player.position_primary or "",
        "club": club_name,
        "league": league_name,
        "step": step,
        "appearances": apps,
        "goals": goals,
    }


def _parse_reasons(match_reasons: dict) -> dict[str, str]:
    """Extract well-known reason keys from the JSONB match_reasons."""
    out: dict[str, str] = {
        "name_similarity": "",
        "same_club": "",
        "same_position": "",
        "dob_match": "",
        "same_nationality": "",
        "other": "",
    }

    if not match_reasons:
        return out

    known_keys = {
        "name_similarity", "name_score", "name",
        "same_club", "club",
        "same_position", "position",
        "dob_match", "dob",
        "same_nationality", "nationality",
    }

    other_parts: list[str] = []

    for key, val in match_reasons.items():
        lk = key.lower().replace(" ", "_")
        if lk in ("name_similarity", "name_score", "name"):
            out["name_similarity"] = str(val)
        elif lk in ("same_club", "club"):
            out["same_club"] = str(val)
        elif lk in ("same_position", "position"):
            out["same_position"] = str(val)
        elif lk in ("dob_match", "dob"):
            out["dob_match"] = str(val)
        elif lk in ("same_nationality", "nationality"):
            out["same_nationality"] = str(val)
        else:
            other_parts.append(f"{key}={val}")

    if other_parts:
        out["other"] = "; ".join(other_parts)

    return out


def export() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Merge Candidates"

    # ── Write headers ────────────────────────────────────────────
    for col_idx, header in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = HEADER_BORDER

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # ── Query pending candidates ─────────────────────────────────
    with get_session() as session:
        candidates = session.execute(
            select(MergeCandidate)
            .where(MergeCandidate.status == "pending")
            .order_by(MergeCandidate.score.desc())
        ).scalars().all()

        log.info("Found %d pending merge candidates", len(candidates))

        if not candidates:
            ws.cell(row=2, column=1, value="No pending merge candidates found.")
            ws.cell(row=2, column=1).font = Font(name="Calibri", italic=True, size=11)
            OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            wb.save(OUTPUT_FILE)
            log.info("Saved empty workbook to %s", OUTPUT_FILE)
            return OUTPUT_FILE

        for row_idx, mc in enumerate(candidates, 2):
            pa = _player_info(session, mc.player_a)
            pb = _player_info(session, mc.player_b)
            reasons = _parse_reasons(mc.match_reasons or {})

            values = [
                mc.id,
                mc.score,
                # Player A
                pa["id"],
                pa["name"],
                pa["dob"],
                pa["position"],
                pa["club"],
                pa["league"],
                pa["step"],
                pa["appearances"],
                pa["goals"],
                # Player B
                pb["id"],
                pb["name"],
                pb["dob"],
                pb["position"],
                pb["club"],
                pb["league"],
                pb["step"],
                pb["appearances"],
                pb["goals"],
                # Reasons
                reasons["name_similarity"],
                reasons["same_club"],
                reasons["same_position"],
                reasons["dob_match"],
                reasons["same_nationality"],
                reasons["other"],
                # Decision (blank for reviewer)
                "",
            ]

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")

            # Alternate row shading
            if row_idx % 2 == 0:
                for col_idx in range(1, len(COLUMNS) + 1):
                    c = ws.cell(row=row_idx, column=col_idx)
                    if c.fill == PatternFill():
                        c.fill = ALT_ROW_FILL

            # Yellow highlight on DECISION column
            decision_col = len(COLUMNS)
            decision_cell = ws.cell(row=row_idx, column=decision_col)
            decision_cell.fill = DECISION_FILL
            decision_cell.font = DECISION_FONT
            decision_cell.alignment = Alignment(horizontal="center")

    # ── Column widths ────────────────────────────────────────────
    col_widths = {
        "Candidate ID": 13,
        "Score": 8,
        "A: Player ID": 12,
        "A: Name": 25,
        "A: DOB": 12,
        "A: Position": 10,
        "A: Club": 25,
        "A: League": 30,
        "A: Step": 8,
        "A: Appearances": 13,
        "A: Goals": 9,
        "B: Player ID": 12,
        "B: Name": 25,
        "B: DOB": 12,
        "B: Position": 10,
        "B: Club": 25,
        "B: League": 30,
        "B: Step": 8,
        "B: Appearances": 13,
        "B: Goals": 9,
        "Name Similarity": 16,
        "Same Club": 11,
        "Same Position": 14,
        "DOB Match": 11,
        "Same Nationality": 16,
        "Other Reasons": 25,
        "DECISION": 12,
    }
    for col_idx, header in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(header, 14)

    # ── Save ─────────────────────────────────────────────────────
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    log.info(
        "Exported %d merge candidates to %s",
        len(candidates), OUTPUT_FILE,
    )
    return OUTPUT_FILE


if __name__ == "__main__":
    path = export()
    print(f"\nDone — {path}")
